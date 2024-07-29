import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile


def carregar_planilha(file, skiprows):
    try:
        planilha = pd.read_excel(file, skiprows=skiprows)
        return planilha
    except Exception as e:
        st.error(f"Erro ao carregar o arquivo: {e}")
        return None


def carregar_todas_abas(file):
    try:
        xls = pd.ExcelFile(file)
        df = pd.concat(
            pd.read_excel(xls, sheet_name=sheet) for sheet in xls.sheet_names
        )
        return df
    except Exception as e:
        st.error(f"Erro ao carregar o arquivo com múltiplas abas: {e}")
        return None


def estilizar_dataframe(df, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Adicionar cabeçalho
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Estilo do cabeçalho
    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border

    # Estilo das células
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border

    return wb


def to_excel_bytes(wb):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def criar_arquivo_zip(arquivos):
    buffer_zip = BytesIO()
    with zipfile.ZipFile(buffer_zip, "w") as zipf:
        for nome_arquivo, dados in arquivos:
            zipf.writestr(nome_arquivo, dados.getvalue())
    buffer_zip.seek(0)
    return buffer_zip


def gerar_planilha_conferencia(df, nome):
    df_conferencia = df.copy()
    df_conferencia["Contagem 1"] = df_conferencia["Contagem"]
    df_conferencia["Contagem 2"] = df_conferencia["Contagem"]
    df_conferencia["Contagem 3"] = df_conferencia["Contagem"]
    df_conferencia["Valor Adotado"] = df_conferencia["Contagem"]
    df_conferencia = df_conferencia[
        [
            "Medicamento",
            "Lote",
            "Data Vencimento",
            "Contagem 1",
            "Contagem 2",
            "Contagem 3",
            "Valor Adotado",
        ]
    ]
    return estilizar_dataframe(df_conferencia, nome)


def main():
    st.title("Sistema de Inventário")

    opcao = st.selectbox(
        "Escolha uma opção:", ["Gerar lista de contagem", "Gerar apuração"]
    )

    if opcao == "Gerar lista de contagem":
        st.subheader("Gerar Lista de Contagem")

        item_selecionado = st.text_input("Nome da Lista:")

        estoque_file = st.file_uploader("Upload da planilha de Estoque:", type=["xlsx"])
        enderecos_file = st.file_uploader(
            "Upload da planilha de Endereços:", type=["xlsx"]
        )

        if estoque_file and enderecos_file:
            estoque_df = carregar_planilha(estoque_file, skiprows=7)
            enderecos_df = carregar_todas_abas(enderecos_file)

            if estoque_df is not None and enderecos_df is not None:

                # Gerando nome das planilhas
                data_atual = datetime.now().strftime("%Y%m%d")
                nome_arquivo_estoque = f"Estoque_{item_selecionado}_{data_atual}.xlsx"
                # Estilizando dataframes
                wb_estoque = Workbook()
                ws_estoque = wb_estoque.active
                ws_estoque.title = "Estoque"
                for r in dataframe_to_rows(estoque_df, index=False, header=True):
                    ws_estoque.append(r)
                for cell in ws_estoque["1:1"]:
                    cell.font = Font(bold=True)
                excel_bytes_estoque = to_excel_bytes(wb_estoque)

                if "Contagem" not in estoque_df.columns:
                    estoque_df["Contagem"] = None

                estoque_df = estoque_df[
                    ["Medicamento", "Lote", "Data Vencimento", "Contagem"]
                ]
                estoque_df["Lote"] = estoque_df["Lote"].astype(str)

                enderecos_df = enderecos_df.rename(
                    columns={
                        "LOCALIZAÇÃO": "Endereço",
                        "PROGRAMA": "Programa",
                        "LOTE": "Lote",
                    }
                )
                enderecos_df["Endereço"].fillna(enderecos_df["05/26"], inplace=True)
                enderecos_df["Lote"] = enderecos_df["Lote"].astype(str).str.rstrip()
                enderecos_df = enderecos_df[["Endereço", "Programa", "Lote"]]

                merged_df = pd.merge(estoque_df, enderecos_df, on="Lote", how="left")
                colunas_reordenadas = [
                    "Endereço",
                    "Medicamento",
                    "Lote",
                    "Data Vencimento",
                    "Programa",
                    "Contagem",
                ]
                merged_df = merged_df[colunas_reordenadas]

                merged_df = merged_df.sort_values(by="Medicamento")

                nome_arquivo_1 = f"{item_selecionado}_contagem1_{data_atual}.xlsx"
                nome_arquivo_2 = f"{item_selecionado}_contagem2_{data_atual}.xlsx"
                nome_arquivo_conferencia = (
                    f"{item_selecionado}_{data_atual}_conferencia.xlsx"
                )

                wb1 = estilizar_dataframe(merged_df, "Contagem 1")
                wb2 = estilizar_dataframe(merged_df, "Contagem 2")
                wb_conferencia = gerar_planilha_conferencia(merged_df, "Conferência")

                # Convertendo para bytes
                excel_bytes_1 = to_excel_bytes(wb1)
                excel_bytes_2 = to_excel_bytes(wb2)
                excel_bytes_conferencia = to_excel_bytes(wb_conferencia)

                # Criando arquivo ZIP
                arquivos = [
                    (nome_arquivo_estoque, excel_bytes_estoque),
                    (nome_arquivo_1, excel_bytes_1),
                    (nome_arquivo_2, excel_bytes_2),
                    (nome_arquivo_conferencia, excel_bytes_conferencia),
                ]
                arquivo_zip_bytes = criar_arquivo_zip(arquivos)

                # Exibir tabelas resultantes
                st.write("Resultado da Análise:")
                st.dataframe(merged_df)

                # Botão de download
                st.download_button(
                    label="Baixar Todos os Arquivos",
                    data=arquivo_zip_bytes,
                    file_name=f"{item_selecionado}_{data_atual}_arquivos.zip",
                    mime="application/zip",
                )

    if opcao == "Gerar apuração":
        st.subheader("Gerar Apuração")
        estoque_file2 = st.file_uploader("Upload da planilha de Estoque:", type=["xlsx"])
        conferencia_file = st.file_uploader("Upload da planilha de Conferencia:", type=["xlsx"])

        if estoque_file2 and conferencia_file:
            conferencia_df = carregar_planilha(conferencia_file, skiprows=0)
            conferencia_df = conferencia_df[
                ["Medicamento", "Lote", "Data Vencimento", "Valor Adotado"]
            ]
            conferencia_df = (
                conferencia_df.groupby(["Medicamento", "Lote", "Data Vencimento"])[
                    "Valor Adotado"
                ]
                .sum()
                .reset_index()
            )
            estoque_df = carregar_planilha(estoque_file2, skiprows=7)
            estoque_df = estoque_df[
                [
                    "Código Simpas",
                    "Medicamento",
                    "Lote",
                    "Data Vencimento",
                    "Quantidade Encontrada",
                    "Valor Unitário",
                    "Programa Saúde",
                ]
            ]
            estoque_df["Lote"] = estoque_df["Lote"].astype(str)
            estoque_df["Código Simpas"] = estoque_df["Código Simpas"].astype(str)
            estoque_df = (
                estoque_df.groupby(
                    [
                        "Código Simpas",
                        "Medicamento",
                        "Lote",
                        "Data Vencimento",
                        "Valor Unitário",
                        "Programa Saúde",
                    ]
                )["Quantidade Encontrada"]
                .sum()
                .reset_index()
            )

            df = pd.merge(
                conferencia_df,
                estoque_df,
                how="left",
                on=["Lote", "Medicamento", "Data Vencimento"],
            )

            df = df.rename(
                columns={
                    "Data Vencimento": "Validade",
                    "Quantidade Encontrada": "SIGAF",
                    "Valor Adotado": "Contagem",
                }
            )

            df['Contagem'] = pd.to_numeric(df['Contagem'], errors='coerce')
            df['SIGAF'] = pd.to_numeric(df['SIGAF'], errors='coerce')
            df['Valor Unitário'] = pd.to_numeric(df['Valor Unitário'], errors='coerce')

            df["Dirença"] = df["Contagem"] - df["SIGAF"]
            df["Vlr Total"] = df["Contagem"] * df["Valor Unitário"]
            df["Vlr Divergencia"] = df["Dirença"] * df["Valor Unitário"]

            new = [
                "Código Simpas",
                "Medicamento",
                "Lote",
                "Validade",
                "Contagem",
                "SIGAF",
                "Dirença",
                "Valor Unitário",
                "Vlr Total",
                "Vlr Divergencia",
                "Programa Saúde",
            ]
            df = df[new]

            # Exibir tabelas resultantes
            st.write("Resultado da Análise:")
            st.dataframe(df)
            
            # Botão de download
                st.download_button(
                    label="Planilha de Conferencia",
                    data=df,
                    file_name=f"Conferencia_{data_atual}.xlsx",
                    mime="application/zip",
                )


if __name__ == "__main__":
    main()
